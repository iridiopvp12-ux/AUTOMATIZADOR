import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def generate_fiscal_report(df: pd.DataFrame, output_path: str):
    """
    Generates an Excel report from the aggregated SPED data.
    """
    if df is None or df.empty:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidação CST_CFOP"

    # Write Main Header (Row 1)
    ws.merge_cells('A1:M1')
    ws['A1'] = "REGISTROS FISCAIS - CONSOLIDAÇÃO DAS OPERAÇÕES POR BLOCO, CFOP E CST"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column Mapping:
    # A: Bloco
    # B: CFOP
    # C: Valor Item
    # D: Valor ICMS
    # E: Valor IPI
    # F: CST (PIS)
    # G: Base PIS
    # H: Aliq PIS
    # I: Valor PIS
    # J: CST (COFINS)
    # K: Base COFINS
    # L: Aliq COFINS
    # M: Valor COFINS

    # PIS Header: F, G, H, I
    ws.merge_cells('F2:I2')
    ws['F2'] = "PIS/PASEP"
    ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F2'].font = Font(bold=True)

    # COFINS Header: J, K, L, M
    ws.merge_cells('J2:M2')
    ws['J2'] = "COFINS"
    ws['J2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['J2'].font = Font(bold=True)

    # Sub Headers (Row 3)
    cols = [
        "Bloco", "CFOP", "Vl Contábil", "Vl ICMS", "Vl IPI",
        "CST", "Base Calc", "Aliq %", "Valor",
        "CST", "Base Calc", "Aliq %", "Valor"
    ]
    for i, col in enumerate(cols):
        cell = ws.cell(row=3, column=i+1, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # Add border
        cell.border = Border(bottom=Side(style='thin'))

    # Data (Row 4+)
    for r_idx, row in df.iterrows():
        # A - B
        ws.cell(row=r_idx+4, column=1, value=row['Bloco'])
        ws.cell(row=r_idx+4, column=2, value=row['CFOP'])

        # C - E (Values)
        ws.cell(row=r_idx+4, column=3, value=row['Valor_Item']).number_format = '#,##0.00'
        ws.cell(row=r_idx+4, column=4, value=row['Valor_ICMS']).number_format = '#,##0.00'
        ws.cell(row=r_idx+4, column=5, value=row['Valor_IPI']).number_format = '#,##0.00'

        # F - I (PIS)
        ws.cell(row=r_idx+4, column=6, value=row['CST_PIS'])
        ws.cell(row=r_idx+4, column=7, value=row['Base_PIS']).number_format = '#,##0.00'
        ws.cell(row=r_idx+4, column=8, value=row['Aliq_PIS']).number_format = '0.00'
        ws.cell(row=r_idx+4, column=9, value=row['Valor_PIS']).number_format = '#,##0.00'

        # J - M (COFINS)
        ws.cell(row=r_idx+4, column=10, value=row['CST_COFINS'])
        ws.cell(row=r_idx+4, column=11, value=row['Base_COFINS']).number_format = '#,##0.00'
        ws.cell(row=r_idx+4, column=12, value=row['Aliq_COFINS']).number_format = '0.00'
        ws.cell(row=r_idx+4, column=13, value=row['Valor_COFINS']).number_format = '#,##0.00'

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter

    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                # Skip checking merged cells content if they are not the top-left
                if isinstance(cell,  pd.DataFrame): # safety check? No, cell is openpyxl cell
                    pass
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return True
